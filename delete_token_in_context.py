import argparse
import json
import os
from string import ascii_uppercase

# pip install uncertainty-calibration
import calibration as cal
import numpy as np
import scipy
import torch
import torch.nn as nn
import transformers
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sklearn.metrics import accuracy_score, f1_score
from tensorboardX import SummaryWriter
from torch import Tensor
from torch.nn import functional as F
from torch.nn.modules.loss import CrossEntropyLoss
from torch.optim.adamw import AdamW
from torch.utils.data import DataLoader, Dataset, RandomSampler
from torch.utils.data.dataloader import DataLoader
from tqdm import tqdm
from transformers import BertConfig, BertForNextSentencePrediction, BertTokenizer

from preprocess_dataset import get_dd_corpus, get_dd_multiref_testset
from utils import (
    RankerDataset,
    dump_config,
    get_uttr_token,
    load_model,
    set_random_seed,
    write2tensorboard,
)

ascii_uppercase = list(ascii_uppercase)
column_titles = (
    ascii_uppercase
    + ["A" + c for c in ascii_uppercase]
    + ["B" + c for c in ascii_uppercase]
    + ["C" + c for c in ascii_uppercase]
    + ["D" + c for c in ascii_uppercase]
    + ["E" + c for c in ascii_uppercase]
)


def visualize():
    wb = Workbook()
    ws = wb.active

    with open("saver_unk.jsonl", "r") as f:
        ls = [json.loads(el) for el in f.readlines()]
    for idx, item in enumerate(ls):
        context, response, original, chnaged = (
            item["context"],
            item["response"],
            item["original_score"],
            item["changed_score"],
        )

        ws["A" + str(idx * 5 + 1)] = "Context"
        ws["A" + str(idx * 5 + 1)].font = Font(bold=True)
        for word_idx, word in enumerate(context):
            ws[column_titles[1 + word_idx] + str(idx * 5 + 1)] = context[word_idx]
            change = round(original - chnaged[word_idx], 2)
            if change < 0:
                ws[column_titles[1 + word_idx] + str(idx * 5 + 2)] = change
                ws[column_titles[1 + word_idx] + str(idx * 5 + 2)].font = Font(
                    bold=True
                )
                ws[column_titles[1 + word_idx] + str(idx * 5 + 2)].fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid"
                )
        ws["A" + str(idx * 5 + 3)] = "Response"
        ws["A" + str(idx * 5 + 3)].font = Font(bold=True)
        for word_idx, word in enumerate(response):
            ws[column_titles[1 + word_idx] + str(idx * 5 + 3)] = response[word_idx]

        ws["A" + str(idx * 5 + 4)] = "Original"
        ws["A" + str(idx * 5 + 4)].font = Font(bold=True)
        ws["B" + str(idx * 5 + 4)] = round(original, 2)
    wb.save("unking.xlsx")


def main(args):
    set_random_seed(42)

    dump_config(args)
    device = torch.device("cuda")
    tokenizer = BertTokenizer.from_pretrained("bert-base-uncased")
    UTTR_TOKEN = get_uttr_token()
    special_tokens_dict = {"additional_special_tokens": [UTTR_TOKEN]}
    tokenizer.add_special_tokens(special_tokens_dict)

    model = BertForNextSentencePrediction.from_pretrained("bert-base-uncased")
    model.resize_token_embeddings(len(tokenizer))
    model = load_model(model, args.model_path, 0, len(tokenizer))
    model.to(device)

    raw_dd_train = get_dd_corpus("train")
    train_dataset = RankerDataset(raw_dd_train, tokenizer, "train", 128, UTTR_TOKEN)
    softmax = torch.nn.Softmax(dim=1)

    saver = []

    for idx in tqdm(range(len(train_dataset))):
        if idx == 1000:
            break
        changed_prediction_list = []
        with torch.no_grad():
            sample = [el[idx] for el in train_dataset.feature]
            # 1 for positive and 0 for random
            label = int(sample[2].numpy())
            if label == 0:
                continue
            ids, masks = [torch.unsqueeze(el, 0).to(device) for el in sample[:2]]
            original_prediction = softmax(model(ids, masks)[0]).cpu().numpy()[0][1]

            for token_index, token_id in enumerate(ids[0]):
                if token_id == tokenizer.sep_token_id:
                    break
                if token_id == tokenizer.cls_token_id:
                    assert token_index == 0
                    continue
                changed_ids = torch.tensor(ids)
                changed_ids[0][token_index] = tokenizer.unk_token_id  # mask_token_id
                changed_prediction = (
                    softmax(model(changed_ids, masks)[0]).cpu().numpy()[0][1]
                )
                changed_prediction_list.append(float(changed_prediction))

        decoded = tokenizer.convert_ids_to_tokens(ids[0])
        if "[PAD]" in decoded:
            decoded = decoded[: decoded.index("[PAD]")]
        decoded = decoded[1:-1]
        context, response = (
            decoded[: decoded.index("[SEP]")],
            decoded[decoded.index("[SEP]") + 1 :],
        )
        assert len(context) == len(changed_prediction_list)

        saver.append(
            {
                "idx": idx,
                "context": context,
                "response": response,
                "original_score": float(original_prediction),
                "changed_score": changed_prediction_list,
            }
        )
    with open("saver_unk.jsonl", "w") as f:
        for l in saver:
            f.write(json.dumps(l))
            f.write("\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process some integers.")
    parser.add_argument(
        "--exp_name",
        type=str,
        default="deterministic_bert_ranker",
    )
    parser.add_argument("--log_path", type=str, default="logs")
    parser.add_argument("--epoch", type=int, default=0)

    args = parser.parse_args()
    args.exp_path = os.path.join(args.log_path, args.exp_name)
    args.model_path = os.path.join(args.exp_path, "model")
    args.board_path = os.path.join(args.exp_path, "board")
    main(args)
    visualize()
